import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json
import os
import time
from datetime import datetime
import re
import io
import hashlib
from typing import Dict, List, Any, Tuple, Optional
from auth import authenticate_user, render_user_header
from utils import (
    safe_unicode_text,
    check_mistral_ocr_availability,
    extract_text_with_mistral_ocr,
    get_api_key,
    get_mistral_api_key,
    get_mistral_api_key_with_session,
    upload_file_to_mistral,
    get_mistral_ocr_result,
)

# Import dependencies with error handling
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import requests
    MISTRAL_AVAILABLE = True
except ImportError:
    MISTRAL_AVAILABLE = False

try:
    from docx import Document
    from docx.shared import Inches
    from docx.enum.text import WD_ALIGN_PARAGRAPH
    from docx.oxml.shared import OxmlElement, qn
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import Color, red, orange, yellow, lightgrey, black
    from reportlab.lib.units import inch
    from reportlab.platypus.flowables import KeepTogether
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import PyPDF2
    import pdfplumber
    PDF_EXTRACT_AVAILABLE = True
except ImportError:
    PDF_EXTRACT_AVAILABLE = False

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Configuration
MAX_CHARS_PER_CHUNK = 100000  # For OpenAI processing
CHUNK_DELAY = 0.5
MAX_RETRIES = 3

# ── File upload validation ────────────────────────────────────────────────────

MAX_FILE_SIZE_MB = 50
ALLOWED_EXTENSIONS = {".pdf", ".docx", ".doc"}
ALLOWED_MIME_TYPES = {
    "application/pdf",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/msword",
}


def validate_uploaded_file(uploaded_file) -> tuple:
    """
    Validate an uploaded file before processing.
    Returns (is_valid: bool, error_message: str).
    """
    if uploaded_file is None:
        return False, "No file uploaded."

    file_size_mb = len(uploaded_file.getvalue()) / (1024 * 1024)
    if file_size_mb > MAX_FILE_SIZE_MB:
        return False, f"File too large ({file_size_mb:.1f} MB). Max size is {MAX_FILE_SIZE_MB} MB."

    ext = os.path.splitext(uploaded_file.name)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        return False, f"Unsupported file type: {ext}. Allowed: {', '.join(sorted(ALLOWED_EXTENSIONS))}"

    if hasattr(uploaded_file, "type") and uploaded_file.type:
        if uploaded_file.type not in ALLOWED_MIME_TYPES:
            return False, f"Invalid file type detected: {uploaded_file.type}"

    return True, ""


# Streamlit App Configuration
st.set_page_config(
    page_title="Film Script Production Breakdown",
    page_icon="🎬",
    layout="wide"
)


# Film Script Processing Classes
class FilmScriptProcessor:
    def __init__(self, openai_api_key: str):
        """Initialize the processor with OpenAI API key"""
        self.client = OpenAI(api_key=openai_api_key)
        self.system_prompt = """You are a senior film production coordinator with extensive experience in script breakdowns for pre-production. Your job is to read screenplay chunks and extract **clear, structured production elements** including locations, scenes, and props."""
        
        self.user_prompt = """You will now be given a screenplay chunk. Based on it, extract the following elements in a **structured JSON format**:

1. **Location-Based Scene Breakdown**
For each LOCATION found in this chunk, group all scenes that occur there:
   * `location_name` (primary key - e.g., "Office", "Restaurant", "Car")
   * `scenes_in_location` (array of scenes at this location):
     - `scene_number` (if available or infer from order)
     - `scene_heading` (e.g., INT. OFFICE – DAY)
     - `time_of_day` (DAY/NIGHT/DAWN/DUSK)
     - `brief_description` (1–2 line summary of scene events)
     - `props_in_scene` (all physical props mentioned explicitly)

2. **Unique Props List**
At the end, give a consolidated list of all **unique props** used across all scenes in this chunk.

Return the response in this exact JSON format:
{
  "location_breakdown": [
    {
      "location_name": "Office",
      "scenes_in_location": [
        {
          "scene_number": 1,
          "scene_heading": "INT. OFFICE - DAY",
          "time_of_day": "DAY",
          "brief_description": "Character enters office and sits at desk",
          "props_in_scene": ["desk", "chair", "computer"]
        }
      ]
    }
  ],
  "unique_props": ["desk", "chair", "computer"]
}

IMPORTANT: Only process the content in this specific chunk. Don't make assumptions about other parts of the script."""

    def extract_text_from_pdf(self, file_data: bytes) -> str:
        """Extract text from PDF bytes"""
        try:
            text = ""
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_data))
            for i, page in enumerate(pdf_reader.pages):
                page_text = page.extract_text()
                text += page_text + "\n"
            return text
        except Exception as e:
            raise Exception(f"Error reading PDF: {str(e)}")

    def extract_text_from_docx(self, file_data: bytes) -> str:
        """Extract text from DOCX bytes"""
        try:
            doc = Document(io.BytesIO(file_data))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise Exception(f"Error reading DOCX: {str(e)}")

    def extract_text_from_file(self, file_data: bytes, filename: str) -> str:
        """Extract text from file based on extension"""
        file_extension = os.path.splitext(filename)[1].lower()
        
        if file_extension == '.pdf':
            return self.extract_text_from_pdf(file_data)
        elif file_extension in ['.docx', '.doc']:
            return self.extract_text_from_docx(file_data)
        else:
            raise ValueError(f"Unsupported file type: {file_extension}")

    def estimate_tokens(self, text: str) -> int:
        """Estimate token count (rough approximation: 1 token ≈ 4 characters)"""
        return len(text) // 4
    
    def chunk_text_for_processing(self, text: str, max_tokens: int = 6000) -> List[str]:
        """Split text into chunks that fit within token limits"""
        # Account for system prompt and response tokens
        available_tokens = max_tokens - 1500  # Reserve tokens for system prompt and response
        max_chars = available_tokens * 4  # Rough conversion
        
        if len(text) <= max_chars:
            return [text]
        
        chunks = []
        lines = text.split('\n')
        current_chunk = ""
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Check if adding this line would exceed the limit
            if len(current_chunk + line + '\n') > max_chars and current_chunk:
                chunks.append(current_chunk.strip())
                current_chunk = line + '\n'
            else:
                current_chunk += line + '\n'
        
        # Add remaining chunk
        if current_chunk.strip():
            chunks.append(current_chunk.strip())
        
        return chunks
    
    def process_chunk_with_openai(self, chunk: str, chunk_num: int, total_chunks: int) -> Dict[str, Any]:
        """Process a single chunk with OpenAI"""
        try:
            # Simplified and more reliable prompt for chunked processing
            chunk_prompt = f"""Analyze this screenplay chunk and extract locations, scenes, and props. Return ONLY a JSON object with no additional text.

Chunk {chunk_num} of {total_chunks}:

{chunk}

Return exactly this JSON structure (no markdown, no explanations):
{{
  "locations": [
    {{
      "location_name": "Location Name",
      "scenes": [
        {{
          "scene_number": "1",
          "scene_heading": "INT. LOCATION - DAY",
          "time_of_day": "DAY",
          "description": "What happens in this scene",
          "props": ["prop1", "prop2"]
        }}
      ]
    }}
  ]
}}"""
            
            response = self.client.chat.completions.create(
                model="gpt-4",
                messages=[
                    {"role": "system", "content": "You are a screenplay analyzer. Extract locations, scenes, and props from screenplay text. Return ONLY valid JSON with no markdown formatting or additional text."},
                    {"role": "user", "content": chunk_prompt}
                ],
                temperature=0.1,
                max_tokens=1500
            )
            
            response_text = response.choices[0].message.content.strip()
            
            # Clean up the response - remove markdown formatting if present
            if response_text.startswith("```json"):
                response_text = response_text.replace("```json", "")
            if response_text.endswith("```"):
                response_text = response_text.replace("```", "")
            
            response_text = response_text.strip()
            
            # Multiple strategies to find JSON
            json_data = None
            
            # Strategy 1: Try direct parsing
            try:
                json_data = json.loads(response_text)
            except json.JSONDecodeError:
                pass
            
            # Strategy 2: Find JSON boundaries
            if not json_data:
                try:
                    start_idx = response_text.find('{')
                    end_idx = response_text.rfind('}') + 1
                    
                    if start_idx != -1 and end_idx > start_idx:
                        json_str = response_text[start_idx:end_idx]
                        json_data = json.loads(json_str)
                except json.JSONDecodeError:
                    pass
            
            # Strategy 3: Look for specific patterns
            if not json_data:
                try:
                    # Try to find JSON-like structures
                    import re
                    json_pattern = r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}'
                    matches = re.findall(json_pattern, response_text, re.DOTALL)
                    
                    for match in matches:
                        try:
                            json_data = json.loads(match)
                            break
                        except json.JSONDecodeError:
                            continue
                except:
                    pass
            
            if json_data:
                return json_data
            else:
                # If no JSON found, create a minimal structure
                return {
                    "locations": [{
                        "location_name": f"Chunk_{chunk_num}_Content",
                        "scenes": [{
                            "scene_number": f"Chunk_{chunk_num}",
                            "scene_heading": f"CONTENT FROM CHUNK {chunk_num}",
                            "time_of_day": "UNKNOWN",
                            "description": f"Content from chunk {chunk_num} - manual review needed",
                            "props": []
                        }]
                    }],
                    "parsing_error": f"Could not parse JSON from chunk {chunk_num}",
                    "raw_response": response_text[:500] + "..." if len(response_text) > 500 else response_text
                }
                
        except Exception as e:
            return {
                "error": f"OpenAI error processing chunk {chunk_num}: {str(e)}",
                "locations": [{
                    "location_name": f"Error_Chunk_{chunk_num}",
                    "scenes": [{
                        "scene_number": f"Error_{chunk_num}",
                        "scene_heading": f"ERROR IN CHUNK {chunk_num}",
                        "time_of_day": "UNKNOWN",
                        "description": f"Error processing chunk {chunk_num}: {str(e)}",
                        "props": []
                    }]
                }]
            }
    
    def merge_chunk_results(self, chunk_results: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Merge results from multiple chunks into final format"""
        merged_locations = {}
        all_props = set()
        processing_errors = []
        
        for i, chunk_result in enumerate(chunk_results, 1):
            if 'error' in chunk_result:
                processing_errors.append(f"Chunk {i}: {chunk_result['error']}")
                continue
            
            if 'parsing_error' in chunk_result:
                processing_errors.append(f"Chunk {i}: {chunk_result['parsing_error']}")
                # Still try to process the fallback structure
            
            locations = chunk_result.get('locations', [])
            for location in locations:
                location_name = location.get('location_name', 'Unknown')
                scenes = location.get('scenes', [])
                
                # Skip error placeholder locations
                if location_name.startswith('Error_Chunk_') or location_name.startswith('Chunk_') and '_Content' in location_name:
                    continue
                
                if location_name not in merged_locations:
                    merged_locations[location_name] = {
                        'location_name': location_name,
                        'scenes_in_location': []
                    }
                
                for scene in scenes:
                    # Skip error placeholder scenes
                    if scene.get('scene_heading', '').startswith('ERROR IN CHUNK') or scene.get('scene_heading', '').startswith('CONTENT FROM CHUNK'):
                        continue
                    
                    # Add scene to location
                    scene_data = {
                        'scene_number': scene.get('scene_number', 'N/A'),
                        'scene_heading': scene.get('scene_heading', 'N/A'),
                        'time_of_day': scene.get('time_of_day', 'N/A'),
                        'brief_description': scene.get('description', 'N/A'),
                        'props_in_scene': scene.get('props', [])
                    }
                    merged_locations[location_name]['scenes_in_location'].append(scene_data)
                    
                    # Collect all props
                    all_props.update(scene.get('props', []))
        
        # Convert to final format
        final_result = {
            'location_breakdown': list(merged_locations.values()),
            'unique_props': sorted(list(all_props))
        }
        
        # Add processing summary
        total_chunks = len(chunk_results)
        successful_chunks = total_chunks - len(processing_errors)
        
        final_result['processing_summary'] = {
            'total_chunks': total_chunks,
            'successful_chunks': successful_chunks,
            'failed_chunks': len(processing_errors),
            'success_rate': f"{(successful_chunks/total_chunks*100):.1f}%" if total_chunks > 0 else "0%"
        }
        
        if processing_errors:
            final_result['processing_errors'] = processing_errors
        
        return final_result
    def process_with_openai(self, script_text: str, progress_callback=None) -> Dict[str, Any]:
        """Process script text with OpenAI using chunking for large texts"""
        try:
            if progress_callback:
                progress_callback(0.1, "🔤 Analyzing script length and preparing for processing...")
            
            # Estimate tokens
            estimated_tokens = self.estimate_tokens(script_text)
            
            if progress_callback:
                progress_callback(0.15, f"📊 Estimated tokens: {estimated_tokens:,}")
            
            # Check if we need to chunk the text
            if estimated_tokens > 6000:  # Safe limit for GPT-4
                if progress_callback:
                    progress_callback(0.2, f"📄 Text is large ({estimated_tokens:,} tokens), splitting into chunks...")
                
                chunks = self.chunk_text_for_processing(script_text)
                
                if progress_callback:
                    progress_callback(0.25, f"📋 Created {len(chunks)} chunks for processing...")
                
                # Process each chunk
                chunk_results = []
                for i, chunk in enumerate(chunks):
                    if progress_callback:
                        progress = 0.3 + (0.6 * (i + 1) / len(chunks))
                        progress_callback(progress, f"🤖 Processing chunk {i+1}/{len(chunks)} with OpenAI...")
                    
                    chunk_result = self.process_chunk_with_openai(chunk, i+1, len(chunks))
                    chunk_results.append(chunk_result)
                    
                    # Small delay to avoid rate limiting
                    time.sleep(0.5)
                
                if progress_callback:
                    progress_callback(0.9, "🔄 Merging results from all chunks...")
                
                # Merge chunk results
                final_result = self.merge_chunk_results(chunk_results)
                
                if progress_callback:
                    progress_callback(0.95, "✅ Successfully processed all chunks!")
                
                return final_result
            
            else:
                # Process as single chunk (original logic with improved JSON handling)
                if progress_callback:
                    progress_callback(0.2, f"📤 Sending {len(script_text):,} characters to OpenAI GPT-4...")
                
                response = self.client.chat.completions.create(
                    model="gpt-4",
                    messages=[
                        {"role": "system", "content": "You are a screenplay analyzer. Extract locations, scenes, and props from screenplay text. Return ONLY valid JSON with no markdown formatting or additional text."},
                        {"role": "user", "content": f"""{self.user_prompt}

Script content:
{script_text}

Return exactly this JSON structure (no markdown, no explanations):"""}
                    ],
                    temperature=0.1,
                    max_tokens=4000
                )
                
                if progress_callback:
                    progress_callback(0.7, "🤖 Received AI response, processing results...")
                
                response_text = response.choices[0].message.content.strip()
                
                if progress_callback:
                    progress_callback(0.8, "📋 Parsing structured data from AI response...")
                
                # Enhanced JSON parsing with multiple strategies
                json_data = None
                
                # Clean up markdown formatting
                if response_text.startswith("```json"):
                    response_text = response_text.replace("```json", "")
                if response_text.endswith("```"):
                    response_text = response_text.replace("```", "")
                response_text = response_text.strip()
                
                # Strategy 1: Direct parsing
                try:
                    json_data = json.loads(response_text)
                except json.JSONDecodeError:
                    pass
                
                # Strategy 2: Find JSON boundaries
                if not json_data:
                    try:
                        start_idx = response_text.find('{')
                        end_idx = response_text.rfind('}') + 1
                        
                        if start_idx != -1 and end_idx > start_idx:
                            json_str = response_text[start_idx:end_idx]
                            json_data = json.loads(json_str)
                    except json.JSONDecodeError:
                        pass
                
                # Strategy 3: Pattern matching
                if not json_data:
                    try:
                        import re
                        json_pattern = r'\{.*\}'
                        match = re.search(json_pattern, response_text, re.DOTALL)
                        if match:
                            json_data = json.loads(match.group())
                    except:
                        pass
                
                if json_data:
                    if progress_callback:
                        progress_callback(0.9, "✅ Successfully parsed production breakdown data...")
                    return json_data
                else:
                    return {
                        "error": "Could not parse JSON from OpenAI response",
                        "raw_response": response_text,
                        "debug_info": f"Response length: {len(response_text)}, First 500 chars: {response_text[:500]}"
                    }
                
        except Exception as e:
            if "context_length_exceeded" in str(e):
                return {
                    "error": f"Script too long for processing. Estimated {self.estimate_tokens(script_text):,} tokens. Please try a shorter script or contact support.",
                    "raw_response": str(e)
                }
            else:
                raise Exception(f"OpenAI processing error: {str(e)}")

    def process_script_file(self, file_data: bytes, filename: str, progress_callback=None) -> Dict[str, Any]:
        """Main processing function - extract text and analyze with OpenAI"""
        try:
            if progress_callback:
                progress_callback(0.0, f"📄 Starting processing of {filename}...")
            
            # Extract text from file
            if progress_callback:
                progress_callback(0.1, f"🔍 Extracting text from {filename}...")
            
            script_text = self.extract_text_from_file(file_data, filename)
            
            if not script_text.strip():
                return {
                    "error": "No text extracted from file",
                    "filename": filename
                }
            
            if progress_callback:
                progress_callback(0.2, f"✅ Extracted {len(script_text):,} characters from {filename}")
            
            # Process with OpenAI
            if progress_callback:
                progress_callback(0.3, "🤖 Starting AI analysis...")
            
            result = self.process_with_openai(script_text, progress_callback)
            
            if progress_callback:
                progress_callback(0.95, "📊 Finalizing results and adding metadata...")
            
            # Add metadata
            if 'error' not in result:
                result['metadata'] = {
                    'filename': filename,
                    'file_size': len(file_data),
                    'text_length': len(script_text),
                    'processed_at': time.strftime('%Y-%m-%d %H:%M:%S')
                }
            
            if progress_callback:
                progress_callback(1.0, "🎉 Processing complete!")
            
            return result
            
        except Exception as e:
            return {
                "error": str(e),
                "filename": filename
            }

# Report Generation Functions
def generate_excel_report(results: Dict[str, Any], filename: str) -> bytes:
    """Generate Excel report from processing results"""
    if not EXCEL_AVAILABLE:
        st.error("Excel generation not available. Please install openpyxl.")
        return None
    
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        # Summary sheet
        summary_sheet = wb.create_sheet("SUMMARY")
        summary_sheet['A1'] = "FILM PRODUCTION BREAKDOWN"
        summary_sheet['A1'].font = Font(bold=True, size=16)
        summary_sheet['A1'].alignment = Alignment(horizontal='center')
        summary_sheet.merge_cells('A1:D1')
        
        # Add metadata
        if 'metadata' in results:
            summary_sheet['A3'] = "Source File:"
            summary_sheet['B3'] = results['metadata'].get('filename', 'N/A')
            summary_sheet['A4'] = "File Size:"
            summary_sheet['B4'] = f"{results['metadata'].get('file_size', 0)} bytes"
            summary_sheet['A5'] = "Processing Date:"
            summary_sheet['B5'] = results['metadata'].get('processed_at', 'N/A')
        
        # Production summary
        total_locations = len(results.get('location_breakdown', []))
        total_scenes = sum(len(loc.get('scenes_in_location', [])) for loc in results.get('location_breakdown', []))
        total_props = len(results.get('unique_props', []))
        
        summary_sheet['A7'] = "PRODUCTION SUMMARY"
        summary_sheet['A7'].font = Font(bold=True, size=14)
        summary_sheet['A8'] = "Total Locations:"
        summary_sheet['B8'] = total_locations
        summary_sheet['A9'] = "Total Scenes:"
        summary_sheet['B9'] = total_scenes
        summary_sheet['A10'] = "Total Props:"
        summary_sheet['B10'] = total_props
        
        # Location overview
        summary_sheet['A12'] = "LOCATION OVERVIEW"
        summary_sheet['A12'].font = Font(bold=True, size=12)
        
        headers = ["Location", "Scenes Count", "Props Count", "Day Scenes", "Night Scenes"]
        for col, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=13, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Fill location data
        row = 14
        for location in results.get('location_breakdown', []):
            location_name = location.get('location_name', 'Unknown')
            scenes = location.get('scenes_in_location', [])
            
            location_props = set()
            day_scenes = night_scenes = 0
            
            for scene in scenes:
                location_props.update(scene.get('props_in_scene', []))
                time_of_day = scene.get('time_of_day', '').upper()
                if 'DAY' in time_of_day:
                    day_scenes += 1
                elif 'NIGHT' in time_of_day:
                    night_scenes += 1
            
            summary_sheet[f'A{row}'] = location_name
            summary_sheet[f'B{row}'] = len(scenes)
            summary_sheet[f'C{row}'] = len(location_props)
            summary_sheet[f'D{row}'] = day_scenes
            summary_sheet[f'E{row}'] = night_scenes
            row += 1
        
        # Master Props List sheet
        props_sheet = wb.create_sheet("MASTER PROPS LIST")
        props_sheet['A1'] = "MASTER PROPS LIST"
        props_sheet['A1'].font = Font(bold=True, size=14)
        
        headers = ["#", "Prop Name", "Locations Used", "Total Scenes"]
        for col, header in enumerate(headers, 1):
            cell = props_sheet.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Fill props data
        row = 4
        for i, prop in enumerate(results.get('unique_props', []), 1):
            prop_locations = []
            prop_scenes = 0
            
            for location in results.get('location_breakdown', []):
                for scene in location.get('scenes_in_location', []):
                    if prop in scene.get('props_in_scene', []):
                        if location.get('location_name') not in prop_locations:
                            prop_locations.append(location.get('location_name'))
                        prop_scenes += 1
            
            props_sheet[f'A{row}'] = i
            props_sheet[f'B{row}'] = prop
            props_sheet[f'C{row}'] = ', '.join(prop_locations)
            props_sheet[f'D{row}'] = prop_scenes
            row += 1
        
        # Create sheets for each location
        for location in results.get('location_breakdown', []):
            location_name = location.get('location_name', 'Unknown')
            sheet_name = location_name.replace('/', '_').replace('\\', '_').replace(':', '_')[:31]
            
            location_sheet = wb.create_sheet(sheet_name)
            location_sheet['A1'] = f"LOCATION: {location_name}"
            location_sheet['A1'].font = Font(bold=True, size=16)
            location_sheet['A1'].fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")
            location_sheet.merge_cells('A1:G1')
            
            # Scene headers
            headers = ["Scene #", "Scene Heading", "Time", "Description", "Props", "Props Count", "Notes"]
            for col, header in enumerate(headers, 1):
                cell = location_sheet.cell(row=5, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Fill scene data
            row = 6
            for scene in location.get('scenes_in_location', []):
                props_list = scene.get('props_in_scene', [])
                
                location_sheet[f'A{row}'] = scene.get('scene_number', 'N/A')
                location_sheet[f'B{row}'] = scene.get('scene_heading', 'N/A')
                location_sheet[f'C{row}'] = scene.get('time_of_day', 'N/A')
                location_sheet[f'D{row}'] = scene.get('brief_description', 'N/A')
                location_sheet[f'E{row}'] = ', '.join(props_list)
                location_sheet[f'F{row}'] = len(props_list)
                location_sheet[f'G{row}'] = ""  # Notes column
                row += 1
            
            # Adjust column widths
            location_sheet.column_dimensions['A'].width = 10
            location_sheet.column_dimensions['B'].width = 25
            location_sheet.column_dimensions['C'].width = 12
            location_sheet.column_dimensions['D'].width = 40
            location_sheet.column_dimensions['E'].width = 30
            location_sheet.column_dimensions['F'].width = 12
            location_sheet.column_dimensions['G'].width = 20
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Excel report: {e}")
        return None

def create_visualization_charts(results: Dict[str, Any]):
    """Create visualization charts for the results"""
    if not results.get('location_breakdown'):
        return None, None
    
    # Prepare data for charts
    location_data = []
    for location in results['location_breakdown']:
        location_name = location.get('location_name', 'Unknown')
        scenes = location.get('scenes_in_location', [])
        
        props = set()
        day_scenes = night_scenes = 0
        
        for scene in scenes:
            props.update(scene.get('props_in_scene', []))
            time_of_day = scene.get('time_of_day', '').upper()
            if 'DAY' in time_of_day:
                day_scenes += 1
            elif 'NIGHT' in time_of_day:
                night_scenes += 1
        
        location_data.append({
            'location': location_name,
            'total_scenes': len(scenes),
            'unique_props': len(props),
            'day_scenes': day_scenes,
            'night_scenes': night_scenes
        })
    
    df = pd.DataFrame(location_data)
    
    # Scenes by location chart
    fig1 = px.bar(
        df,
        x='location',
        y='total_scenes',
        title='Scenes by Location',
        labels={'total_scenes': 'Number of Scenes', 'location': 'Location'},
        color='total_scenes',
        color_continuous_scale='Blues'
    )
    fig1.update_layout(xaxis_tickangle=-45)
    
    # Day vs Night scenes chart
    fig2 = px.bar(
        df,
        x='location',
        y=['day_scenes', 'night_scenes'],
        title='Day vs Night Scenes by Location',
        labels={'value': 'Number of Scenes', 'location': 'Location'},
        barmode='group'
    )
    fig2.update_layout(xaxis_tickangle=-45)
    
    return fig1, fig2

def create_mistral_ocr_tab():
    """Create OCR tab with Mistral OCR integration"""
    st.header("🔍 Mistral OCR Analysis")
    st.markdown("**Extract text from script images using Mistral OCR for production breakdown analysis.**")
    
    # Check Mistral OCR availability
    mistral_available, mistral_message = check_mistral_ocr_availability()
    
    if mistral_available:
        st.success("✅ Mistral OCR is ready!")
        
        # Language selection
        language_options = {
            "Bengali + English (Recommended)": "ben+eng",
            "Hindi + English": "hin+eng", 
            "Tamil + English": "tam+eng",
            "Telugu + English": "tel+eng",
            "English Only": "eng",
            "All Indian Languages": "all"
        }
        
        selected_language = st.selectbox(
            "Select OCR Language",
            options=list(language_options.keys()),
            index=0,
            help="Choose the primary language(s) for OCR recognition"
        )
        
        language_code = language_options[selected_language]
        
        uploaded_image = st.file_uploader(
            "Choose a script image file",
            type=['png', 'jpg', 'jpeg', 'bmp', 'tiff', 'webp'],
            help="Upload an image containing script text for OCR extraction"
        )
        
        if uploaded_image is not None:
            st.success(f"✅ Image uploaded: {uploaded_image.name}")
            
            # Show image preview
            if PIL_AVAILABLE:
                try:
                    image = Image.open(uploaded_image)
                    st.image(image, caption="Uploaded Script Image", use_column_width=True)
                except Exception as e:
                    st.warning(f"Could not display image preview: {e}")
            
            if st.button("🔍 Extract Text with Mistral OCR", type="primary"):
                # Create progress bar for OCR
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                def update_progress(value, message):
                    progress_bar.progress(value)
                    status_text.text(message)
                
                try:
                    extracted_text = extract_text_with_mistral_ocr(uploaded_image, language_code, update_progress)
                except Exception as _ocr_err:
                    st.error(f"❌ Mistral OCR failed: {_ocr_err}")
                    extracted_text = ""

                if extracted_text.strip():
                    st.success(f"✅ Extracted {len(extracted_text):,} characters")
                    
                    with st.expander("📄 Extracted Text Preview"):
                        st.text_area("Extracted Text", extracted_text, height=300)
                    
                    # Process extracted text
                    st.header("🤖 Processing Extracted Script")
                    
                    openai_key = get_api_key()
                    if not openai_key:
                        st.error("❌ OpenAI API key required for script processing")
                        return
                    
                    processor = FilmScriptProcessor(openai_key)
                    
                    # Create fake file data for processing
                    fake_file_data = extracted_text.encode('utf-8')
                    
                    # Create progress bar for script processing
                    processing_progress = st.progress(0)
                    processing_status = st.empty()
                    
                    def update_processing_progress(value, message):
                        processing_progress.progress(value)
                        processing_status.text(message)
                    
                    results = processor.process_script_file(fake_file_data, "OCR_Extracted_Script.txt", update_processing_progress)
                    
                    if 'error' in results:
                        st.error(f"❌ Processing error: {results['error']}")
                    else:
                        st.success("🎉 Script processing complete!")
                        display_results(results, "OCR_Extracted_Script")
                
                else:
                    st.error("❌ No text could be extracted from the image")
    else:
        st.error(f"❌ Mistral OCR not available: {mistral_message}")
        
        # Show setup instructions
        with st.expander("🔧 Setup Instructions"):
            st.markdown("""
            **To enable Mistral OCR:**
            1. ✅ Configure your Mistral API key above
            2. ✅ Ensure OCR access in your Mistral subscription  
            3. ✅ Test the connection
            """)

def display_results(results: Dict[str, Any], filename: str):
    """Display processing results with charts and download options"""
    st.header("📊 Production Breakdown Results")
    
    # Processing completion message
    if 'metadata' in results:
        processing_time = results['metadata'].get('processed_at', 'Unknown')
        st.success(f"✅ Processing completed at {processing_time}")
    
    # Show processing summary if available (for chunked processing)
    if 'processing_summary' in results:
        summary = results['processing_summary']
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Chunks", summary['total_chunks'])
        with col2:
            st.metric("Successful Chunks", summary['successful_chunks'])
        with col3:
            st.metric("Failed Chunks", summary['failed_chunks'])
        with col4:
            st.metric("Success Rate", summary['success_rate'])
        
        # Show errors if any
        if 'processing_errors' in results:
            with st.expander("⚠️ Processing Warnings"):
                for error in results['processing_errors']:
                    st.warning(error)
    
    # Summary metrics
    total_locations = len(results.get('location_breakdown', []))
    total_scenes = sum(len(loc.get('scenes_in_location', [])) for loc in results.get('location_breakdown', []))
    total_props = len(results.get('unique_props', []))
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Locations", total_locations)
    with col2:
        st.metric("Total Scenes", total_scenes)
    with col3:
        st.metric("Total Props", total_props)
    
    # Show message if no content was extracted
    if total_locations == 0 and total_scenes == 0:
        st.warning("⚠️ No locations or scenes were extracted from the script. This might indicate:")
        st.markdown("""
        - The script format is not recognized
        - The content doesn't contain standard screenplay elements
        - The text might need manual formatting
        - Try uploading a different format or checking the script structure
        """)
        
        # Still show raw response for debugging
        if 'raw_response' in results:
            with st.expander("🔍 Debug Information"):
                st.text_area("Raw AI Response", results['raw_response'], height=200)
        
        return
    
    # Charts
    st.subheader("📈 Production Analytics")
    fig1, fig2 = create_visualization_charts(results)
    
    if fig1 and fig2:
        col1, col2 = st.columns(2)
        with col1:
            st.plotly_chart(fig1, use_container_width=True)
        with col2:
            st.plotly_chart(fig2, use_container_width=True)
    
    # Location breakdown
    st.subheader("🏢 Location Breakdown")
    for location in results.get('location_breakdown', []):
        location_name = location.get('location_name', 'Unknown')
        scenes = location.get('scenes_in_location', [])
        
        with st.expander(f"📍 {location_name} ({len(scenes)} scenes)"):
            for scene in scenes:
                st.write(f"**Scene {scene.get('scene_number', 'N/A')}**: {scene.get('scene_heading', 'N/A')}")
                st.write(f"*Time*: {scene.get('time_of_day', 'N/A')}")
                st.write(f"*Description*: {scene.get('brief_description', 'N/A')}")
                st.write(f"*Props*: {', '.join(scene.get('props_in_scene', []))}")
                st.divider()
    
    # Props list
    st.subheader("🎭 Master Props List")
    if results.get('unique_props'):
        props_df = pd.DataFrame({
            'Prop': results['unique_props'],
            'Index': range(1, len(results['unique_props']) + 1)
        })
        st.dataframe(props_df, use_container_width=True)
    else:
        st.info("No props were identified in the script.")
    
    # Download reports
    st.subheader("📥 Download Reports")
    
    # Show report generation progress
    if st.button("🔄 Generate Reports", type="secondary"):
        report_progress = st.progress(0)
        report_status = st.empty()
        
        report_status.text("📊 Generating Excel report...")
        report_progress.progress(0.3)
        
        excel_data = generate_excel_report(results, filename)
        
        report_status.text("📄 Preparing JSON export...")
        report_progress.progress(0.6)
        
        json_data = json.dumps(results, indent=2)
        
        report_status.text("✅ Reports ready for download!")
        report_progress.progress(1.0)
        
        time.sleep(1)  # Brief pause for user feedback
        
        # Download buttons
        col1, col2 = st.columns(2)
        
        with col1:
            if excel_data:
                st.download_button(
                    label="📊 Download Excel Report",
                    data=excel_data,
                    file_name=f"{filename}_production_breakdown.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            st.download_button(
                label="📄 Download JSON Data",
                data=json_data,
                file_name=f"{filename}_breakdown.json",
                mime="application/json"
            )
    else:
        # Show quick download buttons without progress
        col1, col2 = st.columns(2)
        
        with col1:
            excel_data = generate_excel_report(results, filename)
            if excel_data:
                st.download_button(
                    label="📊 Download Excel Report",
                    data=excel_data,
                    file_name=f"{filename}_production_breakdown.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        with col2:
            json_data = json.dumps(results, indent=2)
            st.download_button(
                label="📄 Download JSON Data",
                data=json_data,
                file_name=f"{filename}_breakdown.json",
                mime="application/json"
            )

def main():
    """Main application function"""
    # Authentication check
    if not authenticate_user():
        st.stop()
    render_user_header()

    # Initialise OCR session state
    if "ocr_available" not in st.session_state:
        st.session_state.ocr_available = False
    if "ocr_error_message" not in st.session_state:
        st.session_state.ocr_error_message = ""

    # Custom CSS
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(90deg, #ff6b6b, #4ecdc4);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .user-info {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #007bff;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header
    st.markdown("""
    <div class="main-header">
        <h1>🎬 Film Script Production Breakdown</h1>
        <p>Automated Location & Props Analysis Platform</p>
        <p style="font-size: 0.9em; opacity: 0.9;">✅ AI-Powered Processing • 📊 Excel Reports • 🔍 Mistral OCR • 📋 Production Planning</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown(f"""
        <div class="user-info">
            <h3>👤 User Information</h3>
            <p><b>Name:</b> {st.session_state.get('user_name', 'Unknown')}</p>
            <p><b>Email:</b> {st.session_state.get('user_email', 'unknown')}</p>
            <p><b>Role:</b> {'Admin' if st.session_state.get('is_admin', False) else 'User'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        st.header("🔧 System Status")
        
        # System components status
        components = [
            ("OpenAI API", OPENAI_AVAILABLE),
            ("Mistral API", MISTRAL_AVAILABLE),
            ("DOCX Processing", DOCX_AVAILABLE),
            ("PDF Processing", PDF_EXTRACT_AVAILABLE),
            ("Excel Reports", EXCEL_AVAILABLE),
            ("PDF Generation", PDF_AVAILABLE),
            ("Image Processing", PIL_AVAILABLE)
        ]
        
        for name, available in components:
            if available:
                st.success(f"✅ {name}")
            else:
                st.error(f"❌ {name}")
        
        st.divider()
        
        if st.button("🔄 New Analysis", type="secondary"):
            st.rerun()
        
        if st.button("🚪 Logout", type="secondary"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    # API Configuration
    st.header("🔑 API Configuration")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("OpenAI Configuration")
        openai_key = get_api_key()
        if openai_key:
            st.success("✅ OpenAI API Key: Configured")
        else:
            st.warning("⚠️ OpenAI API Key: Not configured")
            openai_input = st.text_input(
                "Enter OpenAI API Key", 
                type="password", 
                help="Required for script processing"
            )
    
    with col2:
        st.subheader("Mistral Configuration")
        mistral_key = get_mistral_api_key()
        if mistral_key:
            st.success("✅ Mistral API Key: Configured")
        else:
            st.warning("⚠️ Mistral API Key: Not configured")
            mistral_input = st.text_input(
                "Enter Mistral API Key", 
                type="password", 
                help="Required for OCR functionality"
            )
            if mistral_input:
                st.session_state.temp_mistral_key = mistral_input
                st.success("✅ Mistral API Key: Temporarily configured")
    
    # Main tabs
    tab1, tab2, tab3 = st.tabs(["📤 Upload Script", "📝 Paste Script", "🔍 Mistral OCR"])
    
    with tab1:
        st.header("📤 Upload Script File")
        st.markdown("**Upload your screenplay file for automated production breakdown analysis.**")
        
        uploaded_file = st.file_uploader(
            "Choose a script file",
            type=['docx', 'pdf'],
            help="Upload a Microsoft Word document (.docx) or PDF file"
        )
        
        if uploaded_file is not None:
            is_valid, error_msg = validate_uploaded_file(uploaded_file)
            if not is_valid:
                st.error(error_msg)
                st.stop()

            st.success(f"✅ File uploaded: {uploaded_file.name}")

            # Show file info
            file_size = len(uploaded_file.getvalue())
            col1, col2 = st.columns(2)
            with col1:
                st.metric("File Size", f"{file_size/1024:.1f} KB")
            with col2:
                st.metric("File Type", uploaded_file.name.split('.')[-1].upper())
            
            openai_key = get_api_key()
            if not openai_key:
                st.error("❌ OpenAI API key required for processing")
                return
            
            if st.button("🔍 Process Script", type="primary"):
                processor = FilmScriptProcessor(openai_key)
                
                # Create progress bar and status
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                def update_progress(value, message):
                    progress_bar.progress(value)
                    status_text.text(message)
                
                # Show file info and estimated time
                file_size = len(uploaded_file.getvalue())
                estimated_tokens = file_size // 4  # Rough estimate
                
                if estimated_tokens > 6000:
                    st.warning(f"⚠️ Large script detected ({estimated_tokens:,} estimated tokens). Will process in chunks.")
                    estimated_time = max(60, estimated_tokens // 100)  # Longer for chunked processing
                else:
                    estimated_time = max(30, file_size // 1024)
                
                st.info(f"⏱️ Estimated processing time: {estimated_time} seconds")
                
                try:
                    results = processor.process_script_file(uploaded_file.getvalue(), uploaded_file.name, update_progress)
                    
                    if 'error' in results:
                        st.error(f"❌ Processing error: {results['error']}")
                        if 'raw_response' in results:
                            with st.expander("Raw Response"):
                                st.text(results['raw_response'])
                    else:
                        st.success("🎉 Script processing complete!")
                        display_results(results, uploaded_file.name)
                        
                except Exception as e:
                    st.error(f"❌ Unexpected error: {str(e)}")
                    st.info("💡 If the script is very long, try splitting it into smaller sections.")
                
                finally:
                    # Clean up progress indicators
                    progress_bar.empty()
                    status_text.empty()
    
    with tab2:
        st.header("📝 Paste Script Text")
        st.markdown("**Paste your screenplay content for automated production breakdown analysis.**")
        
        script_text = st.text_area(
            "Paste your script content here",
            height=300,
            placeholder="INT. OFFICE - DAY\n\nJOHN enters the office and sits at his desk.\n\nJOHN\nTime to get to work.\n\nHe opens his laptop and starts typing..."
        )
        
        if script_text:
            # Show text info
            word_count = len(script_text.split())
            char_count = len(script_text)
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("Character Count", f"{char_count:,}")
            with col2:
                st.metric("Word Count", f"{word_count:,}")
            
            if st.button("🔍 Process Script", type="primary"):
                openai_key = get_api_key()
                if not openai_key:
                    st.error("❌ OpenAI API key required for processing")
                    return
                
                processor = FilmScriptProcessor(openai_key)
                
                # Create progress bar and status
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                def update_progress(value, message):
                    progress_bar.progress(value)
                    status_text.text(message)
                
                # Show text info and estimated time
                estimated_tokens = char_count // 4  # Rough estimate
                
                if estimated_tokens > 6000:
                    st.warning(f"⚠️ Large script detected ({estimated_tokens:,} estimated tokens). Will process in chunks.")
                    estimated_time = max(60, estimated_tokens // 100)  # Longer for chunked processing
                else:
                    estimated_time = max(20, char_count // 5000)
                
                st.info(f"⏱️ Estimated processing time: {estimated_time} seconds")
                
                try:
                    fake_file_data = script_text.encode('utf-8')
                    results = processor.process_script_file(fake_file_data, "Pasted_Script.txt", update_progress)
                    
                    if 'error' in results:
                        st.error(f"❌ Processing error: {results['error']}")
                        if 'raw_response' in results:
                            with st.expander("Raw Response"):
                                st.text(results['raw_response'])
                    else:
                        st.success("🎉 Script processing complete!")
                        display_results(results, "Pasted_Script")
                        
                except Exception as e:
                    st.error(f"❌ Unexpected error: {str(e)}")
                    st.info("💡 If the script is very long, try splitting it into smaller sections.")
                
                finally:
                    # Clean up progress indicators
                    progress_bar.empty()
                    status_text.empty()
    
    with tab3:
        create_mistral_ocr_tab()
    
    # Footer
    st.markdown("---")
    st.markdown(f"""
    <div style='text-align: center; color: #666; font-size: 0.9em;'>
        <p>🎬 Film Script Production Breakdown System | AI-Powered Analysis | Reviewed by: {st.session_state.get('user_name', 'Unknown')}</p>
        <p>🔒 Secure access • 📊 Automated breakdowns • 🔍 OCR support • 📋 Production planning</p>
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
